import sqlite3
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------- CONFIG ----------------
DB_PATH = r"C:\Users\Widya HW\OneDrive - Universitas Tarumanagara\Desktop\Skripsi\Fix Program\bengkel.db"
OUTPUT_DIR = r"C:\Users\Widya HW\OneDrive - Universitas Tarumanagara\Desktop\Skripsi\Fix Program\Kartu_Pekerjaan"
DEFAULT_COL_WIDTH = 12
TITLE_TEXT = "KARTU PEKERJAAN"
# ----------------------------------------


def to_title_case(text: str) -> str:
    if not isinstance(text, str):
        return ""
    return " ".join(w.capitalize() for w in text.strip().split())


def sanitize_text_for_filename(s: str) -> str:
    """Hapus spasi, jadikan huruf besar, hilangkan karakter aneh."""
    s = s.strip().upper()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s or "file"


def build_excel(name, plat, jenis, tanggal_datang, nomor_antrian, keluhan_list, out_path):

    wb = Workbook()
    ws = wb.active

    # Set width A..G
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_COL_WIDTH

    # Title A1:G2
    ws.merge_cells("A1:G2")
    cell_title = ws["A1"]
    cell_title.value = TITLE_TEXT
    cell_title.font = Font(name="Arial", size=20, bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    # Labels
    ws["A3"] = "Tanggal :"
    ws["A4"] = "Nama :"
    ws["A5"] = "Plat No :"
    ws["D3"] = "Mobil :"
    ws["D4"] = "Mekanik :"

    for c in ["A3", "A4", "A5", "D3", "D4"]:
        ws[c].font = Font(name="Arial", size=10)

    # Fill Data
    ws["B3"] = tanggal_datang
    ws["B4"] = to_title_case(name)
    ws["B5"] = to_title_case(plat)
    ws["E3"] = to_title_case(jenis)

    # Header
    ws["A7"] = "Nomor"
    ws["B7"] = "Keluhan"
    ws["E7"] = "Solusi"
    ws.merge_cells("B7:D7")
    ws.merge_cells("E7:G7")

    for c in ["A7", "B7", "E7"]:
        ws[c].font = Font(name="Arial", size=10, bold=True)
        ws[c].alignment = Alignment(horizontal="center", vertical="center")

    start_row = 8
    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write keluhan
    for i, kel in enumerate(keluhan_list, start=1):
        row = start_row + i - 1

        ws[f"A{row}"] = i
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].font = Font(name="Arial", size=10)

        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = to_title_case(kel)
        ws[f"B{row}"].font = Font(name="Arial", size=10)

        ws.merge_cells(f"E{row}:G{row}")
        ws[f"E{row}"].font = Font(name="Arial", size=10)

    # Apply border
    last_row = start_row + len(keluhan_list) - 1
    if last_row < 7:  
        last_row = 7

    for r in range(7, last_row + 1):
        for col_idx in range(1, 8):
            ws[f"{get_column_letter(col_idx)}{r}"].border = border

    wb.save(out_path)
    print("Saved:", out_path)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id_kartu, nama_pengirim, plat_nomor, jenis_mobil, keluhan,
               tanggal_datang, nomor_antrian
        FROM kartu_pekerjaan
        ORDER BY tanggal_datang, nomor_antrian
    """)

    rows = cur.fetchall()
    conn.close()

    for r in rows:
        nama = r["nama_pengirim"]
        plat = r["plat_nomor"]
        jenis = r["jenis_mobil"]
        keluhan = r["keluhan"]
        tanggal_datang = r["tanggal_datang"]
        nomor_antrian = r["nomor_antrian"]

        # Split keluhan by comma
        kel_list = [k.strip() for k in keluhan.split(",") if k.strip()]

        # Build file name
        plat_clean = sanitize_text_for_filename(plat)
        filename = f"{tanggal_datang}_{plat_clean}_{nomor_antrian}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, filename)

        build_excel(
            nama,
            plat,
            jenis,
            tanggal_datang,
            nomor_antrian,
            kel_list,
            out_path,
        )


if __name__ == "__main__":
    main()
